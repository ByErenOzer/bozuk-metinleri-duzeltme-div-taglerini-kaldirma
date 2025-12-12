import re, html
from openpyxl import load_workbook

def clean_text(s):
    if s is None:
        return s
    if not isinstance(s, str):
        s = str(s)
    t = s
    for _ in range(2):
        t2 = html.unescape(t)
        if t2 == t:
            break
        t = t2
    t = re.sub(r'(?i)<br\s*/?>', '\n', t)
    t = re.sub(r'(?i)</\s*(div|p|span)\s*>', '\n', t)
    t = re.sub(r'(?i)<\s*(div|p|span)[^>]*>', '', t)
    t = re.sub(r'(?i)<[^>]+>', '', t)
    t = t.replace('\u00a0', ' ')
    t = re.sub(r'\s+\n', '\n', t)
    t = re.sub(r'\n\s+', '\n', t)
    t = re.sub(r'\n{3,}', '\n\n', t)
    t = re.sub(r'[ \t]{2,}', ' ', t)
    t = t.strip()
    return t

def process_workbook(src, dst):
    wb = load_workbook(src)
    changed = {}
    for ws in wb.worksheets:
        header_row = ws[1]
        target_cols = []
        for cell in header_row:
            if cell.value:
                name = str(cell.value).strip().lower()
                if name in ('tetkik_sonucu','tetkit_sonucu'):
                    target_cols.append(cell.column)
                    cell.value = 'tetkik_sonucu_temiz'
        count = 0
        for col_idx in target_cols:
            for row in range(2, ws.max_row+1):
                c = ws.cell(row=row, column=col_idx)
                before = c.value
                after = clean_text(before)
                c.value = after
                if before != after:
                    count += 1
        if target_cols:
            changed[ws.title] = count
    wb.save(dst)
    print('saved', dst)
    for k,v in changed.items():
        print(f'{k}: {v}')
    return changed

if __name__ == '__main__':
    src = r'c:\\Users\\doganeren.ozer\\Desktop\\div_clean_text\\patolojiLLMguncel.xlsx'
    dst = r'c:\\Users\\doganeren.ozer\\Desktop\\div_clean_text\\patolojiLLMguncel-temiz.xlsx'
    process_workbook(src, dst)